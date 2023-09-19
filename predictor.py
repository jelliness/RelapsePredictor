import os
import openpyxl

wb = openpyxl.load_workbook(r'SmokingDataSet (3).xlsx')
sheet = wb["Attempt2"]

def getDataFromSheet(target):
    extractedData=[]
    for i in range(1, sheet.max_row + 1):
        cell_obj = sheet.cell(row = i, column = 1)
        if str(cell_obj.value)==target:
            for j in range(1, sheet.max_column + 1):
                cell_obj = sheet.cell(row = i, column = j)
                extractedData.append(cell_obj.value)
    return extractedData

def squeeze(target):
    record=[]
    for data in target:
        record.append(getDataFromSheet(data))
    return record

def getValues(record,index):
    allVal=[]
    for group in record:
        allVal.append(group[index])
    return allVal

def calculateLaplace(num,total,a,K):
    laplace=(num+a)/(total+a*K)
    return laplace

def checkZero(list):
    laplaceExist=False
    if 0 in list:
        laplaceExist=True
    else:
        laplaceExist=False
    return laplaceExist

def getFinalValues(allVal,laplaceExist,total,a,K):
    finalVal=[]
    
    if laplaceExist:
        for num in allVal:
            partialNumerator=calculateLaplace(num,total,a,K)
            finalVal.append(partialNumerator)
    else:
        for num in allVal:
            partialNumerator=(num)/(total)
            finalVal.append(partialNumerator)

    return finalVal

def getP_target(allVal,total,overall,a,K):
    laplaceExist=checkZero(allVal)
    p_yes=0
    if laplaceExist:
        p_yes=calculateLaplace(total,overall,a,K)
    else:
        p_yes=(total)/(overall) 
    return p_yes

def getNumerator(record,index,numOfResponse,overall,a,K):
    numerator=1
    rawNum=getValues(record,index)
    laplaceExist=checkZero(rawNum)
    p_target=getP_target(rawNum,numOfResponse,overall,a,K)
    partialNumerator=getFinalValues(rawNum,laplaceExist,numOfResponse,a,K)
    for num in partialNumerator:
        numerator=numerator*num
    product=numerator*p_target
    return product

def getDenominator(record,overall,checkIndex,a,K,i=3):
    denominator=1
    checkList=getValues(record,checkIndex)
    laplaceExist=checkZero(checkList)
    rawDenom=getValues(record,i)
    partialDenominator=getFinalValues(rawDenom,laplaceExist,overall,a,K)
    for num in partialDenominator:
        denominator=denominator*num
    return denominator

def calculateResponse(record,numOfResponse,overallNumOfData,a,K,index):
    predictedValue=0
    numerator=getNumerator(record,index,numOfResponse,overallNumOfData,a,K)
    denominator=getDenominator(record,overallNumOfData,index,a,K,i=3)
    predictedValue=numerator/denominator    
    return predictedValue

def menu():
    target = []
    N_age = ""
    user_input = input("\nHow old are you?: ")
    if int(user_input) < 17:
        N_age = "CHILDREN"
    elif 16 < int(user_input) < 31:
        N_age = "YOUNG ADULTS"
    elif 30 < int(user_input) < 46:
        N_age = "MIDDLE ADULTS"
    elif int(user_input) > 45:
        N_age = "OLD ADULTS"
    target.append(N_age)

    N_gender = "M"
    user_input = input("Gender [M or F]: ")
    if user_input == "M":
        N_gender = "M"
    else:
        N_gender = "F"
    target.append(N_gender)

    N_cvStatus = ""
    print("\n[1] - Single\n[2] - Married\n")
    user_input = input("Civil Status: ")
    if user_input == "1":
        N_cvStatus = "Single"
    else:
        N_cvStatus = "Married"
    target.append(N_cvStatus)

    N_cessation = ""
    user_input = input("\nDo you have info cessation [Y/N]: ")
    if user_input == "Y":
        N_cessation = "Yes"
    else:
        N_cessation = "No"
    target.append(N_cessation)

    N_empStatus = ""
    print("\n[1] - Employed\n[2] - NotOfficeing\n[3] - Officeing\n[4] - Retired")
    user_input = input("Employment Status: ")
    if user_input == "1":
        N_empStatus = "Employed"
    elif user_input == "2":
        N_empStatus = "NotOfficeing"
    elif user_input == "3":
        N_empStatus = "Officeing"
    else:
        N_empStatus = "Retired"
    target.append(N_empStatus)

    N_type = ""
    print("\n[1] - Regular Smoker\n[2] - Social Smoker")
    user_input = input("What type of smoker are you?: ")
    if user_input == "1":
        N_type = "RegularSmoker"
    else:
        N_type = "SocialSmoker"
    target.append(N_type)

    N_ageStart=""
    user_input = input("\nHow old were you when you started smoking?: ")
    if int(user_input) < 15:
        N_ageStart = "10 - 14"
    elif 14 < int(user_input) < 20:
        N_ageStart = "15 - 19"
    elif int(user_input) > 20:
        N_ageStart = "Above 20"
    target.append(N_ageStart)

    N_influence=''
    print('\n[1] - Curiosity\n[2] - Family Influence\n[3] - Peer Pressure')
    user_input = input("What factors or influences led you to smoking?: ")
    if user_input == "1":
        N_influence = "Curiosity"
    elif user_input == "2":
        N_influence = "FamilyInfluence"
    else:
        N_influence = "PeerPressure"
    target.append(N_influence)

    N_urge = ""
    print('\n[1] - Stressed\n[2] - Bored\n[3] - Sad\n[4] - Angry\n[5] - Happy')
    user_input = input("What specific situations have led you to increase your urge to smoke? : ")
    if user_input == "1":
        N_urge = "Stressed"
    elif user_input == "2":
        N_urge = "Bored"
    elif user_input == "3":
        N_urge = "Sad"
    elif user_input == "4":
        N_urge = "Angry"
    else:
        N_urge = "Happy"
    target.append(N_urge)

    N_noSticks=''
    user_input = input("\nHow many cigarettes or 'sticks' do you typically smoke in a day? : ")
    if int(user_input) < 6:
        N_noSticks = "1 - 5"
    elif 5 < int(user_input) < 11:
        N_noSticks = "6 - 10"
    elif 10 < int(user_input) < 16:
        N_noSticks = "11 - 15"
    elif 15 < int(user_input) < 21:
        N_noSticks = "16 - 20"
    elif 20 < int(user_input) < 26:
        N_noSticks = "21 - 25"
    elif 25 < int(user_input) < 31:
        N_noSticks = "26 - 30"
    elif 30 < int(user_input) < 36:
        N_noSticks = "31 - 35"
    elif 35 < int(user_input) < 41:
        N_noSticks = "36 - 40"
    target.append(N_noSticks)

    N_mainAccess=''
    print('\n[1] - Home\n[2] - Office\n[3] - Public Place\n[4] - Bars\n[5] - Others')
    user_input = input("Where do you often smoke?: ")
    if user_input == "1":
        N_mainAccess = "Home"
    elif user_input == "2":
        N_mainAccess = "Office"
    elif user_input == "3":
        N_mainAccess = "PublicPlace"
    elif user_input == "4":
        N_mainAccess = "Bars"
    else:
        N_mainAccess = "Others"
    target.append(N_mainAccess)
    return target



#target=['YOUNG ADULTS','M','Single','No','Officeing','RegularSmoker','15 - 19','FamilyInfluence','Sad','6 - 10','Bars']
a=1
K=12
totalOfYesResponse=21
totalOfNoResponse=48
overallNumOfData=totalOfYesResponse+totalOfNoResponse

target=menu()
record=squeeze(target)

RelapsePositive=calculateResponse(record,totalOfYesResponse,overallNumOfData,a,K,index=1)
RelapseNegative=calculateResponse(record,totalOfNoResponse,overallNumOfData,a,K,index=2)

overallTotal=RelapsePositive+RelapseNegative
negativeRelapseRate=RelapseNegative/overallTotal
positiveRelapseRate=RelapsePositive/overallTotal
totalPercentage=positiveRelapseRate+negativeRelapseRate
print("\n\n------------------------------------\n\n")
print("RESULTS\n")
print("Prosterior Probability to Relapse: ",RelapsePositive)
print("Prosterior Probability Not to Relapse: ",RelapseNegative)
print()
print(f'Positive Chance to Relapse %: {round(positiveRelapseRate * 100, 4)}%')
print(f'Negative Chance to Relapse %: {round(negativeRelapseRate * 100, 4)}%')
print(f'Total Percentage: {round(totalPercentage * 100, 4)}%')

if positiveRelapseRate > negativeRelapseRate:
    print("\nREMARKS: The smoker is most likely to relapse.")
else:
    print("\nREMARKS: The smoker is most likely not to relapse.")


